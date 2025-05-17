import openpyxl
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import matplotlib.pyplot as plt

# Define Excel path
excel_path = r"C:\Users\SAINATH\Desktop\financeexcel.xlsx"

# Load or create workbook
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    if "Monthly Finance" not in wb.sheetnames:
        ws = wb.create_sheet("Monthly Finance")
    else:
        ws = wb["Monthly Finance"]
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Finance"

# Add headers if sheet is empty
if ws.max_row == 1 and ws['A1'].value is None:
    ws.append(["Date", "Month", "Type", "Category", "Amount", "Budget Limit"])

# Data containers
transactions = []  # List of tuples: (date, month, type, category, amount)
budget_limits = {}  # Dictionary: {"category": limit}

def input_income(month):
    """Add income with date and source."""
    date = datetime.now().strftime("%Y-%m-%d")
    source = input("Enter income source: ")
    amount = float(input(f"Enter amount from {source}: "))
    ws.append([date, month, "Income", source, amount, ""])
    transactions.append((date, month, "Income", source, amount))
    wb.save(excel_path)
    print(f"Income recorded: {amount} Rs. from {source}")

def set_budget_limits(month):
    """Set multiple budget categories at once."""
    date = datetime.now().strftime("%Y-%m-%d")
    num_categories = int(input("How many budget categories do you want to set? "))
    for _ in range(num_categories):
        category = input("Enter category name: ")
        limit = float(input(f"Enter budget limit for {category}: "))
        budget_limits[category] = limit
        ws.append([date, month, "Budget", category, "", limit])
    print(f"\nBudget limits set for {list(budget_limits.keys())}")

def add_expenses(month):
    """Add multiple expenses under budget categories."""
    date = datetime.now().strftime("%Y-%m-%d")
    print("\nAvailable categories:", list(budget_limits.keys()))
    category = input("Enter category to add expenses to: ")
    if category not in budget_limits:
        print("Error: Category not in budget!")
        return

    num_expenses = int(input(f"How many expenses for {category}? "))
    for _ in range(num_expenses):
        name = input("Enter expense description: ")
        amount = float(input("Enter expense amount: "))
        ws.append([date, month, "Expense", category, amount, ""])
        transactions.append((date, month, "Expense", category, amount))
        print(f"Added expense: {amount} Rs. for {name} under {category}")
    wb.save(excel_path)

def generate_report(month):
    """Generate detailed spending report."""
    total_income = sum(t[4] for t in transactions if t[2] == "Income" and t[1] == month)
    total_expenses = sum(t[4] for t in transactions if t[2] == "Expense" and t[1] == month)
    remaining = total_income - total_expenses

    print(f"\n--- {month.upper()} REPORT ---")
    print(f"Total Income: {total_income} Rs.")
    print(f"Total Expenses: {total_expenses} Rs.")
    print(f"Remaining Balance: {remaining} Rs.")

    # Category-wise breakdown
    print("\nCategory-wise Spending:")
    for category in budget_limits:
        spent = sum(t[4] for t in transactions if t[3] == category and t[1] == month)
        limit = budget_limits[category]
        print(f"{category}: {spent} / {limit} Rs.")
        if spent > limit:
            print(f"  ⚠ Overspent by {spent - limit} Rs.")

def show_pie_chart(month):
    """Generate pie chart for expenses."""
    categories = {}
    for t in transactions:
        if t[1] == month and t[2] == "Expense":
            categories[t[3]] = categories.get(t[3], 0) + t[4]
    
    if categories:
        plt.pie(categories.values(), labels=categories.keys(), autopct="%1.1f%%")
        plt.title(f"Spending Breakdown for {month}")
        plt.show()
    else:
        print("No expenses to display.")

def main():
    """Menu-driven interface."""
    while True:
        print("\n----- PERSONAL FINANCE TRACKER -----\n")
        print("1. Add Income")
        print("2. Set Budget Limits (Multiple Categories)")
        print("3. Add Expenses (Multiple Entries)")
        print("4. Generate Report")
        print("5. View Pie Chart")
        print("6. Exit")
        
        choice = input("Enter choice (1-6): ")
        month = input("Enter month (e.g., January): ").strip().capitalize()

        if choice == "1":
            input_income(month)
        elif choice == "2":
            set_budget_limits(month)
        elif choice == "3":
            add_expenses(month)
        elif choice == "4":
            generate_report(month)
        elif choice == "5":
            show_pie_chart(month)
        elif choice == "6":
            print("Exiting...")
            break
        else:
            print("Invalid choice.")

if __name__ == "__main__":
    main()
